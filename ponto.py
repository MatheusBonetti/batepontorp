import discord
from discord.ext import commands
import datetime
import asyncio
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
import os

# Define os intents que seu bot usará
# Certifique-se de que estes intents estão ativados no Portal de Desenvolvedores do Discord
intents = discord.Intents.default()
intents.message_content = True
intents.members = True # **CRÍTICO:** Necessário para bot.get_guild().get_member() para pegar display_name

# Crie uma instância do bot
bot = commands.Bot(command_prefix='/', intents=intents, case_insensitive=True)

# Estrutura para armazenar os pontos em memória
# Cada entrada para um usuário terá:
# {
#   'inicio_geral': datetime (quando o ponto foi iniciado pela primeira vez)
#   'ultimo_inicio_periodo': datetime (quando o último período de trabalho ativo começou/recomeçou)
#   'tempo_trabalhado_total': timedelta (tempo total que o ponto ficou efetivamente ativo)
#   'status': 'ativo' ou 'pausado' (apenas em memória, não persistido no XLSX)
#   'mensagem_id': id da mensagem do ponto para atualizar (apenas em memória)
#   'canal_id': id do canal onde a mensagem do ponto está (apenas em memória)
# }
pontos_usuarios = {}

# Nome do arquivo da planilha .xlsx
EXCEL_FILE_NAME = 'registros_pontos.xlsx'

# Definir os cabeçalhos da planilha na ordem desejada para o registro FINALIZADO
PLANILHA_HEADERS = [
    'NOME_USUARIO',           # Apelido/Display Name do usuário no servidor
    'DATA_REGISTRO',          # Data em que o ponto foi finalizado
    'HORARIO_INICIO_EXPEDIENTE', # Horário do início do expediente
    'HORARIO_FIM_EXPEDIENTE',    # Horário do fim do expediente
    'TEMPO_TOTAL_TRABALHADO_DIA' # Tempo líquido no dia, descontando pausas
]

# --- Funções para manipulação do arquivo .xlsx ---

def _get_excel_sheet():
    """
    Retorna o workbook e a sheet ativa. Se o arquivo não existir, cria um novo
    com os cabeçalhos.
    """
    try:
        workbook = load_workbook(EXCEL_FILE_NAME)
        sheet = workbook.active
        # Verifica se os cabeçalhos estão presentes e adiciona se não estiverem
        current_headers = [sheet.cell(row=1, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
        if current_headers != PLANILHA_HEADERS:
            print("Cabeçalhos da planilha desatualizados ou ausentes. Recriando cabeçalhos.")
            # Ao recriar, perderá dados antigos. Faça backup se precisar.
            new_workbook = Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = "Registros de Pontos"
            new_sheet.append(PLANILHA_HEADERS)
            new_workbook.save(EXCEL_FILE_NAME)
            workbook = new_workbook
            sheet = new_sheet
            print(f"Planilha '{EXCEL_FILE_NAME}' atualizada com novos cabeçalhos. Dados antigos podem ter sido perdidos.")
    except FileNotFoundError:
        # Cria um novo workbook e sheet se o arquivo não existir
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Registros de Pontos"
        sheet.append(PLANILHA_HEADERS) # Adiciona os cabeçalhos
        workbook.save(EXCEL_FILE_NAME)
    except Exception as e:
        print(f"Erro ao carregar ou inicializar planilha '{EXCEL_FILE_NAME}': {e}")
        raise # Propaga o erro para que a execução pare ou seja tratada em nível superior
    return workbook, sheet

def salvar_dados_finalizados(user_id: str):
    """
    Salva os dados de um ponto FINALIZADO na planilha, adicionando uma NOVA LINHA.
    Esta função é chamada apenas quando um ponto é finalizado.
    """
    workbook, sheet = _get_excel_sheet()
    
    user_data = pontos_usuarios.get(user_id)
    if not user_data:
        print(f"Erro ao tentar finalizar ponto para {user_id}: Dados não encontrados na memória.")
        return

    # --- LÓGICA ATUALIZADA PARA PEGAR O NOME DE EXIBIÇÃO/APELIDO ---
    nome_usuario = f"Usuário ID: {user_id}" # Fallback padrão

    # Tenta obter o objeto do canal onde o ponto foi iniciado
    channel = bot.get_channel(user_data['canal_id'])
    
    if channel and isinstance(channel, discord.TextChannel) and channel.guild:
        # Se o canal for um canal de texto e pertencer a uma guilda, tenta pegar o Member
        member_obj = channel.guild.get_member(int(user_id))
        if member_obj:
            nome_usuario = member_obj.display_name # Pega o apelido ou nome de exibição
        else:
            # Se não conseguiu obter o Member (usuário saiu do servidor, cache inválido, etc.)
            user_obj = bot.get_user(int(user_id))
            if user_obj:
                nome_usuario = user_obj.name # Usa o nome de usuário global
    else:
        # Se não conseguiu obter o canal ou a guilda, tenta pegar o nome de usuário global
        user_obj = bot.get_user(int(user_id))
        if user_obj:
            nome_usuario = user_obj.name
    # --- FIM DA LÓGICA ATUALIZADA ---
    
    # Coleta os dados para o registro final
    data_registro = datetime.datetime.now().strftime("%d/%m/%Y") # Data da finalização do ponto
    horario_inicio_expediente = user_data['inicio_geral'].strftime("%H:%M:%S")
    horario_fim_expediente = datetime.datetime.now().strftime("%H:%M:%S")
    tempo_trabalhado_final = user_data.get('tempo_trabalhado_total', datetime.timedelta(seconds=0))
    tempo_trabalhado_str = _formatar_tempo_total(tempo_trabalhado_final)

    # Prepara a NOVA LINHA com os dados finais, COM 'NOME_USUARIO' NA PRIMEIRA POSIÇÃO
    new_row_values = [
        nome_usuario,              # Agora é a primeira coluna
        data_registro,
        horario_inicio_expediente,
        horario_fim_expediente,
        tempo_trabalhado_str
    ]
    
    sheet.append(new_row_values) # Adiciona a nova linha
    
    try:
        workbook.save(EXCEL_FILE_NAME)
        print(f"Ponto do usuário '{nome_usuario}' finalizado e adicionado a '{EXCEL_FILE_NAME}'.")
    except Exception as e:
        print(f"Erro ao salvar planilha '{EXCEL_FILE_NAME}' durante finalização para {user_id}: {e}")
        print("Certifique-se de que o arquivo não está aberto em outro programa (ex: Excel).")
        
    # Remove o ponto da memória após a finalização bem-sucedida
    if user_id in pontos_usuarios:
        del pontos_usuarios[user_id]


def carregar_dados():
    """
    Apenas inicializa o arquivo Excel se não existir e garante os cabeçalhos.
    NÃO carrega pontos ativos/pausados para a memória, pois eles não são persistidos.
    """
    global pontos_usuarios
    pontos_usuarios = {} # Garante que o dicionário de pontos ativos/pausados está vazio ao iniciar
    _get_excel_sheet() # Chama para garantir que o arquivo e os cabeçalhos existam
    print(f"Planilha '{EXCEL_FILE_NAME}' verificada e pronta para registros. Pontos ativos não são restaurados ao iniciar.")


# --- Funções Auxiliares ---
def _formatar_tempo_total(delta_tempo: datetime.timedelta) -> str:
    """
    Formata um objeto timedelta em uma string legível (ex: "2 horas, 30 minutos").
    """
    total_seconds = int(delta_tempo.total_seconds())
    horas = total_seconds // 3600
    minutos = (total_seconds % 3600) // 60
    segundos = total_seconds % 60

    tempo_formatado = []
    if horas > 0:
        tempo_formatado.append(f'{horas} hora{"s" if horas > 1 else ""}')
    if minutos > 0:
        tempo_formatado.append(f'{minutos} minuto{"s" if minutos > 1 else ""}')
    if segundos > 0 or not tempo_formatado: # Garante que "0 segundos" seja mostrado se for o caso
        tempo_formatado.append(f'{segundos} segundo{"s" if segundos > 1 else ""}')
    return ", ".join(tempo_formatado)

def _criar_embed_ponto(user: discord.Member, inicio: datetime.datetime, status: str) -> discord.Embed:
    """
    Cria um objeto Embed para exibir o status do ponto.
    """
    embed = discord.Embed(title="Bate-Ponto", color=0x00ff00)
    embed.add_field(name="Usuário:", value=user.mention, inline=False)
    embed.add_field(name="Início:", value=f'{inicio.strftime("%d de %B de %Y %H:%M")}', inline=False)
    if status == 'pausado':
        embed.description = "Ponto Pausado"
        embed.color = 0xffa500 # Laranja
    embed.set_footer(text="Sistema desenvolvido por Matheus Bonetti - Versão completa.")
    return embed

# --- Views para os botões ---
class PontoView(discord.ui.View):
    """
    View para o ponto ATIVO, com botões de Pausar e Finalizar.
    """
    def __init__(self, user_id: str, timeout=None):
        super().__init__(timeout=timeout)
        self.user_id = user_id

    @discord.ui.button(label="Pausar", style=discord.ButtonStyle.blurple, custom_id="pausar_ponto")
    async def pausar_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        if str(interaction.user.id) != self.user_id:
            await interaction.response.send_message("Você não pode interagir com o ponto de outro usuário.", ephemeral=True)
            return

        user_data = pontos_usuarios.get(self.user_id)
        if not user_data or user_data.get('status') == 'pausado':
            await interaction.response.send_message("Seu ponto já está pausado ou não existe.", ephemeral=True)
            return

        # Calcula o tempo trabalhado no último período ativo
        if 'tempo_trabalhado_total' not in user_data:
            user_data['tempo_trabalhado_total'] = datetime.timedelta(seconds=0)

        tempo_trabalhado_neste_periodo = datetime.datetime.now() - user_data['ultimo_inicio_periodo']
        user_data['tempo_trabalhado_total'] += tempo_trabalhado_neste_periodo
        user_data['ultimo_inicio_periodo'] = datetime.datetime.now() # Marca o tempo da pausa
        user_data['status'] = 'pausado'
        
        # Não há salvamento no Excel aqui, apenas atualização da memória do bot

        embed = _criar_embed_ponto(interaction.user, user_data['inicio_geral'], 'pausado')
        embed.add_field(name="Status:", value="Pausado", inline=False)
        await interaction.response.edit_message(embed=embed, view=PausaView(self.user_id))

    @discord.ui.button(label="Finalizar expediente", style=discord.ButtonStyle.red, custom_id="terminar_ponto")
    async def terminar_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        if str(interaction.user.id) != self.user_id:
            await interaction.response.send_message("Você não pode interagir com o ponto de outro usuário.", ephemeral=True)
            return

        user_data = pontos_usuarios.get(self.user_id)
        if not user_data:
            await interaction.response.send_message("Você não tem um ponto em aberto para finalizar.", ephemeral=True)
            return

        # Se o ponto estiver ativo, calcula o tempo do último período antes de finalizar
        if user_data.get('status') == 'ativo':
            if 'tempo_trabalhado_total' not in user_data:
                user_data['tempo_trabalhado_total'] = datetime.timedelta(seconds=0)
            tempo_trabalhado_neste_periodo = datetime.datetime.now() - user_data['ultimo_inicio_periodo']
            user_data['tempo_trabalhado_total'] += tempo_trabalhado_neste_periodo

        # Salva o registro finalizado na planilha (nova linha) e remove da memória
        # Coleta os dados necessários ANTES de chamar salvar_dados_finalizados,
        # pois ela irá deletar user_data da memória.
        horario_inicio_expediente = user_data['inicio_geral'].strftime("%H:%M:%S")
        horario_fim_expediente = datetime.datetime.now().strftime("%H:%M:%S")
        tempo_trabalhado_final = user_data.get('tempo_trabalhado_total', datetime.timedelta(seconds=0))
        tempo_trabalhado_str = _formatar_tempo_total(tempo_trabalhado_final)

        await interaction.response.send_message(f"Seu ponto foi finalizado. Registrando na planilha...", ephemeral=True) # Feedback imediato

        salvar_dados_finalizados(user_id=self.user_id) # Agora pode salvar e deletar da memória

        # Atualiza a mensagem no Discord (pode ser um follow-up)
        embed = discord.Embed(title="Bate-Ponto", color=0x2ecc71)
        embed.add_field(name="Usuário:", value=interaction.user.mention, inline=False)
        embed.add_field(name="Início do Expediente:", value=horario_inicio_expediente, inline=False)
        embed.add_field(name="Fim do Expediente:", value=horario_fim_expediente, inline=False)
        embed.add_field(name="Tempo total trabalhado:", value=tempo_trabalhado_str, inline=False)
        embed.set_footer(text="Sistema desenvolvido por Matheus Bonetti - Versão completa.")
        # Como já houve um response.send_message, precisa usar follow-up ou edit_original_response
        await interaction.message.edit(embed=embed, view=None)


class PausaView(discord.ui.View):
    """
    View para o ponto PAUSADO, com botões de Continuar e Finalizar (Pausado).
    """
    def __init__(self, user_id: str, timeout=None):
        super().__init__(timeout=timeout)
        self.user_id = user_id

    @discord.ui.button(label="Continuar", style=discord.ButtonStyle.green, custom_id="continuar_ponto")
    async def continuar_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        if str(interaction.user.id) != self.user_id:
            await interaction.response.send_message("Você não pode interagir com o ponto de outro usuário.", ephemeral=True)
            return

        user_data = pontos_usuarios.get(self.user_id)
        if not user_data or user_data.get('status') == 'ativo':
            await interaction.response.send_message("Seu ponto já está ativo ou não existe.", ephemeral=True)
            return

        user_data['ultimo_inicio_periodo'] = datetime.datetime.now()
        user_data['status'] = 'ativo'
        
        # Não há salvamento no Excel aqui, apenas atualização da memória do bot

        embed = _criar_embed_ponto(interaction.user, user_data['inicio_geral'], 'ativo')
        embed.add_field(name="Status:", value="Ativo", inline=False)
        await interaction.response.edit_message(embed=embed, view=PontoView(self.user_id))

    @discord.ui.button(label="Finalizar serviço (Pausado)", style=discord.ButtonStyle.red, custom_id="terminar_ponto_pausado")
    async def terminar_pausado_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        # Reusa a lógica de terminar do PontoView, que já lida com o cálculo do tempo total
        # e a atualização da planilha.
        await PontoView(self.user_id).terminar_button(interaction, button)


# --- Eventos do Bot ---
@bot.event
async def on_ready():
    """
    Evento que é executado quando o bot está pronto e conectado ao Discord.
    Apenas inicializa o arquivo Excel. Pontos ativos/pausados não são restaurados após restart.
    """
    print(f'Bot logado como {bot.user}')
    
    # Apenas verifica/cria o arquivo Excel com os cabeçalhos.
    # Não carrega dados de pontos ativos/pausados para a memória, pois eles não são persistidos.
    carregar_dados() 

    # Não há views para recarregar para pontos ativos, pois eles não são restaurados após restart.
    print("Nenhum ponto ativo será restaurado após o reinício do bot.")


# --- Comandos do Bot ---
@bot.command(name='ponto', aliases=['entrar', 'qap', 'trabalhar', 'lida', 'clt', 'nãoquerotrabalhar'])
async def iniciar_ponto(ctx: commands.Context):
    """
    Comando para iniciar um novo ponto de trabalho.
    """
    user_id = str(ctx.author.id)

    if user_id in pontos_usuarios:
        await ctx.send(f'{ctx.author.mention}, você já tem um ponto em aberto! Use os botões na sua mensagem de ponto para interagir com ele.', ephemeral=True)
        return

    agora = datetime.datetime.now()
    pontos_usuarios[user_id] = {
        'inicio_geral': agora,
        'ultimo_inicio_periodo': agora,
        'tempo_trabalhado_total': datetime.timedelta(seconds=0),
        'status': 'ativo',
        'mensagem_id': None, # Será preenchido após o envio da mensagem
        'canal_id': ctx.channel.id # Salva o ID do canal para recuperar o Member depois
    }
    
    # Não há salvamento no Excel aqui. O ponto só será registrado ao ser FINALIZADO.
    # Os dados estão apenas na memória RAM do bot enquanto ativo/pausado.

    embed = _criar_embed_ponto(ctx.author, agora, 'ativo')
    embed.add_field(name="Status:", value="Ativo", inline=False)

    view = PontoView(user_id)
    message = await ctx.send(embed=embed, view=view)

    # Armazena o ID da mensagem e do canal para que o bot possa atualizá-la.
    # Isso é apenas em memória. Se o bot reiniciar, essa associação será perdida.
    pontos_usuarios[user_id]['mensagem_id'] = message.id
    pontos_usuarios[user_id]['canal_id'] = ctx.channel.id


# **IMPORTANTE**: Substitua 'COLE_O_SEU_TOKEN_REAL_AQUI' pelo token do seu bot!
# Você pode obter seu token em: https://discord.com/developers/applications
bot.run('MTM5NjU5MTUxMTUzMjk5ODY3Ng.GJ1SM3.ba8s_mQVMp5hm9f5B3PrrbndQXAyz6cwIpmg0E') # Substitua este token!